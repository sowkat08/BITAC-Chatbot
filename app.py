import os
import json
import traceback
import re
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from PyPDF2 import PdfReader
import docx
import openpyxl
import requests
from groq import Groq
import numpy as np
import urllib3

# SSL ওয়ার্নিং বন্ধ করার জন্য
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

load_dotenv()

# [CRITICAL CONFIG] Render ড্যাশবোর্ডে শুধু GROQ_API_KEY এবং PORT সেট করবেন
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
PORT = int(os.getenv("PORT", 10000))

# ডিফল্ট ব্যাকআপ ডাটা
BITAC_FALLBACK_INFO = """
বাংলাদেশ Industrial Technical Assistance Center (BITAC - বিটাক) শিল্প मंत्रालয়ের অধীন একটি সরকারি স্বায়ত্তশাসিত কারিগরি প্রতিষ্ঠান।
প্রধান কাজ: শিল্প ক্ষেত্রে উৎপাদনশীলতা বৃদ্ধি, কারিগরি সহায়তা প্রদান, এবং খুচরা যন্ত্রপাতি (Spare Parts) তৈরি করা।
এখানে বিভিন্ন মেয়াদি কারিগরি ও ইন্ডাস্ট্রিয়াল প্রশিক্ষণ প্রদান করা হয় (যেমন: machine shop, ওয়েল্ডিং, ইলেকট্রিক্যাল, অটোমোবাইল, ক্যাড/ক্যাম)।
"""

# ভেক্টর ডাটাবেজ মেমোরি হোল্ডার
CHUNKS_TEXTS = []
CHUNKS_EMBEDDINGS = []

def get_embedding(text: str):
    """[FIXED] Hugging Face বাদ দিয়ে সরাসরি Groq API দিয়ে ফ্রিতে এমবেডিং তৈরি (০ এমবি র‍্যাম খরচ)"""
    if not GROQ_API_KEY:
        print("⚠️ GROQ_API_KEY পাওয়া যায়নি! মক ভেক্টর রিটার্ন করা হচ্ছে।")
        return [0.0] * 1024  # nomic-embed-text-v1.5 এর ডাইমেনশন ১০২৪
        
    try:
        groq_client = Groq(api_key=GROQ_API_KEY)
        # Groq এর অফিশিয়াল এবং অত্যন্ত ফাস্ট এমবেডিং মডেল
        response = groq_client.embeddings.create(
            input=[text],
            model="nomic-embed-text-v1.5"
        )
        return response.data[0].embedding
    except Exception as e:
        print(f"⚠️ Groq এমবেডিং তৈরিতে সমস্যা: {e}")
        return None

def clean_scraped_text(text: str) -> str:
    """Jina AI থেকে আসা লাইভ ওয়েবসাইটের কোড বা আবর্জনা ক্লিন করার ফাংশน"""
    text = re.sub(r'[\t ]+', ' ', text)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    cleaned_lines = []
    for line in lines:
        if any(keyword in line.lower() for keyword in ["javascript:", "css", "nav", "footer", "copyright", "privacy policy"]):
            continue
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines)

def split_text_into_chunks(text: str, chunk_size=400, overlap=50):
    """বাক্য ও প্যারাগ্রাফের অর্থ বজায় রেখে উন্নত উপায়ে টেক্সট টুকরো করার লজিক"""
    paragraphs = text.split("\n\n")
    chunks = []
    current_chunk = []
    current_length = 0
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        words = para.split()
        if current_length + len(words) <= chunk_size:
            current_chunk.append(para)
            current_length += len(words)
        else:
            if current_chunk:
                chunks.append("\n".join(current_chunk))
            current_chunk = [para]
            current_length = len(words)
            
    if current_chunk:
        chunks.append("\n".join(current_chunk))
        
    final_chunks = []
    for c in chunks:
        words = c.split()
        if len(words) > chunk_size + 100:
            for i in range(0, len(words), chunk_size - overlap):
                final_chunks.append(" ".join(words[i:i + chunk_size]))
        else:
            final_chunks.append(c)
            
    return [c.strip() for c in final_chunks if c.strip()]

def load_all_combined_context():
    """ফাইল ও লিঙ্ক থেকে ডাটা পড়ে ভেক্টর ডাটাবেজ তৈরি করার মাস্টার ফাংশন"""
    global CHUNKS_TEXTS, CHUNKS_EMBEDDINGS
    CHUNKS_TEXTS = []
    CHUNKS_EMBEDDINGS = []
    combined_text = ""
    
    # === অংশ ১: লোকাল ফাইল রিড করা ===
    data_dir = "./data"
    if os.path.exists(data_dir):
        print("📁 'data' ফোল্ডার পাওয়া গেছে। ফাইল স্ক্যান করা হচ্ছে...")
        for file in os.listdir(data_dir):
            file_path = os.path.join(data_dir, file)
            try:
                if file.lower().endswith('.pdf'):
                    print(f"📖 PDF রিড হচ্ছে: {file}")
                    reader = PdfReader(file_path)
                    for page in reader.pages:
                        combined_text += page.extract_text() or ""
                elif file.lower().endswith('.docx'):
                    print(f"📖 DOCX রিড হচ্ছে: {file}")
                    doc = docx.Document(file_path)
                    for para in doc.paragraphs:
                        if para.text.strip():
                            combined_text += para.text + "\n"
                elif file.lower().endswith('.xlsx'):
                    print(f"📖 XLSX রিড হচ্ছে: {file}")
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    for sheet in wb.sheetnames:
                        for row in wb[sheet].iter_rows(values_only=True):
                            row_text = " ".join([str(cell) for cell in row if cell is not None])
                            if row_text.strip():
                                combined_text += row_text + "\n"
                elif file.lower().endswith(('.txt', '.json')):
                    print(f"📖 টেক্সট ফাইল রিড হচ্ছে: {file}")
                    with open(file_path, 'r', encoding='utf-8') as f:
                        combined_text += f.read() + "\n"
            except Exception as e:
                print(f"⚠️ ফাইল {file} পড়তে সমস্যা: {e}")
    else:
        print("ℹ️ 'data' ফোল্ডার পাওয়া যায়নি। শুধু লিঙ্ক স্ক্র্যাপার কার্যকর থাকবে।")
                
    # === অংশ ২: Jina AI দিয়ে লাইভ ওয়েবসাইট লিঙ্ক স্ক্র্যাপ করা ===
    target_urls = [
        "https://www.bitac.gov.bd", 
        "https://www.bitac.gov.bd/site/page/89531ca1-a83d-4c31-9257-8fb6fc9ef444"
    ]
    print("🌐 Jina AI এর মাধ্যমে লাইভ ওয়েবসাইট স্ক্র্যাপ করা হচ্ছে...")
    for url in target_urls:
        try:
            jina_proxy_url = f"https://r.jina.ai/{url}"
            response = requests.get(jina_proxy_url, timeout=15)
            if response.status_code == 200 and len(response.text.strip()) > 100:
                cleaned_scraped = clean_scraped_text(response.text.strip())
                combined_text += "\n" + cleaned_scraped + "\n"
                print(f"✅ লিঙ্ক থেকে ডাটা এসেছে: {url}")
        except Exception as e:
            print(f"⚠️ লিঙ্ক স্ক্র্যাপ করতে সমস্যা: {url} -> {e}")

    final_text = combined_text.strip()
    if len(final_text) < 100:
        final_text = BITAC_FALLBACK_INFO
        print("⚠️ কোনো ডাটা পাওয়া যায়নি, ফলব্যাক ডাটা ব্যবহার করা হচ্ছে।")

    raw_chunks = split_text_into_chunks(final_text)
    print(f"🧠 {len(raw_chunks)} টি প্যারাগ্রাফে ডাটা ভাগ করা হয়েছে। ভেক্টরাইজেশন শুরু হচ্ছে...")
    
    for idx, chunk in enumerate(raw_chunks):
        embedding = get_embedding(chunk)
        if embedding:
            CHUNKS_TEXTS.append(chunk)
            CHUNKS_EMBEDDINGS.append(embedding)
            
    print(f"🎉 সাফল্য! মোট {len(CHUNKS_TEXTS)} টি প্যারাগ্রাফ সফলভাবে ভেক্টর ডাটাবেজে আপলোড হয়েছে।")

def get_semantic_context(query: str, top_k=3):
    """ইউজারের প্রশ্নের আসল ভাবার্থ বুঝে ডাটাবেজ থেকে নিখুঁত ৩টি প্যারাগ্রাফ খুঁজে বের করার ফাংশন"""
    global CHUNKS_TEXTS, CHUNKS_EMBEDDINGS
    if not CHUNKS_EMBEDDINGS:
        return ""

    query_embedding = get_embedding(query)
    if not query_embedding:
        return ""

    query_vec = np.array(query_embedding)
    similarities = []
    
    for doc_vec in CHUNKS_EMBEDDINGS:
        doc_vec = np.array(doc_vec)
        dot_product = np.dot(query_vec, doc_vec)
        norm_q = np.linalg.norm(query_vec)
        norm_d = np.linalg.norm(doc_vec)
        
        if norm_q == 0 or norm_d == 0:
            similarity = 0.0
        else:
            similarity = float(dot_product / (norm_q * norm_d))
        similarities.append(similarity)

    top_indices = np.argsort(similarities)[::-1][:top_k]
    if len(top_indices) == 0:
        return ""
        
    max_score = similarities[top_indices[0]]
    print(f"🔍 Top Match Score (Groq Embed): {max_score:.4f}")

    # [FIXED] Groq Nomic এমবেডিংয়ের জন্য পারফেক্ট কসাইন সিমিলারিটি থ্রেশহোল্ড ০.৩৮ সেট করা হলো
    if max_score < 0.38: 
        return "" 

    relevant_chunks = [CHUNKS_TEXTS[idx] for idx in top_indices]
    return "\n\n".join(relevant_chunks)

# [FIXED] ডেপ্রিকেশন ওয়ার্নিং দূর করতে আধুনিক Lifespan যুক্ত করা হলো
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("⏳ সার্ভার চালু হচ্ছে, ব্যাকগ্রাউন্ড ডাটা প্রসেস শুরু হচ্ছে...")
    try:
        load_all_combined_context()
    except Exception as e:
        print(f"❌ স্টার্টআপ প্রসেসে সমস্যা: {e}")
    yield
    print("⏳ সার্ভার বন্ধ হচ্ছে...")

app = FastAPI(lifespan=lifespan)

# CORS পারমিশন
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.head("/")
async def head_home():
    return HTMLResponse(content="", status_code=200)

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return HTMLResponse(content="", status_code=200)

@app.get("/", response_class=HTMLResponse)
def home():
    html_content = """
    <!DOCTYPE html>
    <html lang="bn">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>BITAC Tech-Bot</title>
        <style>
            body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f0f2f5; margin: 0; padding: 0; display: flex; justify-content: center; align-items: center; height: 100vh; }
            .chat-container { width: 100%; max-width: 450px; height: 85vh; background: white; border-radius: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.1); display: flex; flex-direction: column; overflow: hidden; }
            .chat-header { background: #006643; color: white; padding: 15px; text-align: center; font-weight: bold; font-size: 1.1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            .chat-box { flex: 1; padding: 15px; overflow-y: auto; display: flex; flex-direction: column; gap: 12px; background: #f9f9f9; }
            .message { padding: 10px 14px; border-radius: 8px; max-width: 75%; font-size: 0.95rem; line-height: 1.4; word-wrap: break-word; }
            .user-msg { background: #006643; color: white; align-self: flex-end; border-bottom-right-radius: 2px; }
            .bot-msg { background: #e4e6eb; color: #1c1e21; align-self: flex-start; border-bottom-left-radius: 2px; }
            .input-area { display: flex; padding: 12px; gap: 8px; background: white; border-top: 1px solid #eee; }
            .input-area input { flex: 1; padding: 12px; border: 1px solid #ccd0d5; border-radius: 20px; outline: none; font-size: 0.95rem; padding-left: 15px; }
            .input-area input:focus { border-color: #006643; }
            .input-area button { padding: 10px 20px; background: #006643; color: white; border: none; border-radius: 20px; cursor: pointer; font-weight: bold; transition: background 0.2s; }
            .input-area button:hover { background: #004d32; }
        </style>
    </head>
    <body>
    <div class="chat-container">
        <div class="chat-header">BITAC Tech-Bot (Official)</div>
        <div class="chat-box" id="chatBox">
            <div class="message bot-msg">আসসালামু আলাইকুম! বিটাক (BITAC) হেল্পডেস্কে আপনাকে স্বাগতম। আজ কীভাবে সাহায্য করতে পারি?</div>
        </div>
        <div class="input-area">
            <input type="text" id="userInput" placeholder="আপনার প্রশ্নটি এখানে লিখুন..." onkeypress="handleKeyPress(event)">
            <button onclick="sendMessage()">পাঠান</button>
        </div>
    </div>
    <script>
        const BACKEND_URL = window.location.origin + "/chat"; 
        async function sendMessage() {
            const inputField = document.getElementById("userInput");
            const chatBox = document.getElementById("chatBox");
            const query = inputField.value.trim();
            if (!query) return;
            chatBox.innerHTML += `<div class="message user-msg">${query}</div>`;
            inputField.value = "";
            chatBox.scrollTop = chatBox.scrollHeight;
            const loadingId = "loading-" + Date.now();
            chatBox.innerHTML += `<div class="message bot-msg" id="${loadingId}"><i>উত্তর তৈরি হচ্ছে...</i></div>`;
            chatBox.scrollTop = chatBox.scrollHeight;
            try {
                const response = await fetch(`${BACKEND_URL}?user_question=${encodeURIComponent(query)}`, {
                    method: "POST"
                });
                const data = await response.json();
                document.getElementById(loadingId).remove();
                chatBox.innerHTML += `<div class="message bot-msg">${data.response || "দুঃখিত, কোনো উত্তর পাওয়া যায়নি।"}</div>`;
            } catch (error) {
                document.getElementById(loadingId).remove();
                chatBox.innerHTML += `<div class="message bot-msg" style="color: red;">দুঃখিত, উত্তর তৈরিতে সমস্যা হয়েছে।</div>`;
                console.error("Error:", error);
            }
            chatBox.scrollTop = chatBox.scrollHeight;
        }
        function handleKeyPress(event) {
            if (event.key === 'Enter') { sendMessage(); }
        }
    </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

@app.post("/chat")
async def chat_with_bot(user_question: str):
    if not GROQ_API_KEY:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY পাওয়া যায়নি।")
    
    dynamic_context = get_semantic_context(user_question)
        
    if not dynamic_context:
        # [HARDENED PROMPT] ডাটাবেজে ম্যাচ না করলে অবান্তর বা বানোয়াট উত্তর বন্ধ করার জন্য কঠোর গাইডলাইন
        system_prompt = """
        তুমি বিটাক (BITAC)-এর একজন পেশাদার তথ্য কর্মকর্তা।
        ইউজার এমন একটি প্রশ্ন করেছে যার সুনির্দিষ্ট উত্তর আমাদের অফিশিয়াল ডেটা ফাইলে সরাসরি খুঁজে পাওয়া যায়নি।
        
        কঠোর নিয়মাবলী (Strict Rules):
        ১. ইউজার সাধারণ শুভেচ্ছা জানালে (যেমন: হাই, হ্যালো, কেমন আছেন, আসসালামু আলাইকুম) সংক্ষেপে অত্যন্ত বিনয়ের সাথে শুভেচ্ছা বিনিময় করবে।
        ২. এছাড়া যেকোনো সুনির্দিষ্ট নোটিশ, ট্রেনিং কোর্স, ভর্তি বিজ্ঞপ্তি বা ফি নিয়ে প্রশ্ন হলে যা তোমার মেমরিতে বা রেফারেন্স ফাইলে নেই, নিজের মেমোরি থেকে কোনো তথ্য বা কোর্সের নাম বানাবে বা অনুমান করে বলবে না।
        ৩. অত্যন্ত বিনয়ের সাথে স্পষ্ট বাংলায় বলবে: "দুঃখিত, এই সুনির্দিষ্ট তথ্যটি আমার ডাটাবেজে বা বিটাক ওয়েবসাইটে খুঁজে পাওয়া যায়নি।" কোনো মনগড়া বা কাল্পনিক তথ্য দিয়ে ইউজারকে বিভ্রান্ত করবে না।
        """
    else:
        # [STRICT PROMPT] ডাটাবেজে তথ্য থাকলে টু-দ্য-পয়েন্ট উত্তর সাজানোর গাইডলাইন
        system_prompt = f"""
        তুমি বাংলাদেশ ইন্ডাস্ট্রিয়াল টেকনিক্যাল অ্যাসিস্ট্যান্স সেন্টার (BITAC)-এর একজন অফিসিয়াল তথ্য প্রদানকারী চ্যাটবট।
        তোমার প্রধান দায়িত্ব হলো নিচে দেওয়া 'বিটাক রেফারেন্স ডাটা' ব্যবহার করে ইউজারের প্রশ্নের টু-দ্য-পয়েন্ট সঠিক উত্তর দেওয়া।
        
        কঠোর নিয়মাবলী (Strict Rules):
        ১. শুধু এবং কেবলমাত্র নিচে দেওয়া 'বিটাক রেফারেন্স ডাটা'-র ওপর ভিত্তি করে উত্তর তৈরি করবে।
        ২. রেফারেন্স ডাটায় যে তথ্যের স্পষ্ট উল্লেখ নেই, তা নিয়ে নিজের মেমোরি থেকে কোনো কথা বাড়িয়ে বা বানিয়ে বলবে না। বিশেষ করে কোনো কাল্পনিক কারিগরি বা আইটি কোর্সের নাম মেলাবে না।
        ৩. যদি তথ্যটি রেফারেন্সে না থাকে, তবে সরাসরি বলবে তথ্যটি নেই।
        ４. উত্তর সবসময় স্পষ্ট, প্রফেশনাল এবং বাংলা ভাষায় বুলেট পয়েন্ট আকারে প্রাতিষ্ঠানিক টোনে উপস্থাপন করবে।
        
        বিটাক রেফারেন্স ডাটা:
        {dynamic_context}
        """
    try:
        groq_client = Groq(api_key=GROQ_API_KEY)
        chat_completion = groq_client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_question}
            ],
            model="llama-3.3-70b-versatile", 
            temperature=0.0  # [CRITICAL] বানোয়াট উত্তর প্রতিরোধ করতে ফিক্সড ০.০ তাপমাত্রা
        )
        return {"response": chat_completion.choices[0].message.content}
    except Exception as e:
        print("❌ GROQ API CALL FAILED!")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT)
